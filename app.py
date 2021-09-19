import sys,os,re
import openpyxl
from openpyxl import Workbook
from flask import Flask, flash, request, redirect, render_template, send_file, send_from_directory
from werkzeug.utils import secure_filename

from config_simple import *

from pysqlcipher3 import dbapi2 as sqlite3
from config import app_key, db_loc


app=Flask(__name__)
app.secret_key = app_key
app.config['MAX_CONTENT_LENGTH'] = file_mb_max * 1024 * 1024


# Check that the upload folder exists
if not os.path.isdir(upload_dest):
    os.mkdir(upload_dest)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in extensions
    

## on page load display the upload file
@app.route('/upload')
def upload_form():
    return render_template('upload_simple.html')

def file_process():
    # Archivos de Excel desde donde extraeran los datos
    odoo_bn = '/uploads_folder/bancoNacion.xlsx'
    odoo_dc67 = '/uploads_folder/difCambioCta67.xlsx'
    odoo_dc77 = '/uploads_folder/difCambioCta77.xlsx'

    # Archivo con el formato de la plantilla de salida
    plantilla = '/uploads_folder/Desktop/plantilla.xlsx'

    wb_odoo_bn = openpyxl.load_workbook(odoo_bn)
    wb_odoo_dc67 = openpyxl.load_workbook(odoo_dc67)
    wb_odoo_dc77 = openpyxl.load_workbook(odoo_dc77)
    wb_plantilla = openpyxl.load_workbook(plantilla)

    # TODO:Revisar como se pueden tomar los datos siempre de la primera hoja sin necesidad de ponerle un nombre para referenciarla aquí
    odoo_datos_bn = wb_odoo_bn['Data']
    odoo_datos_dc67 = wb_odoo_dc67['Data']
    odoo_datos_dc77 = wb_odoo_dc77['Data']

    template = wb_plantilla['Data']

    row = 2
    col = 0

    i = 3
    while (str(odoo_datos_bn['E' + str(i + 1)].value) != 'None'):
        i += 1
        # Obtengo la descripción de la operación en el informe del banco de la nación
        codNac = odoo_datos_bn['E' + str(i)]
        print(codNac.value)
        if 'CUST.IN/' in codNac.value:
            # Busco ese código en difCambioCta67
            j = 3
            while (str(odoo_datos_dc67['E' + str(j)].value) != 'None'):
                j += 1
                codCta67 = odoo_datos_dc67['E' + str(j)]
                if codNac.value == codCta67.value:
                    template.cell(row=row, column=col + 1, value=odoo_datos_bn['D' + str(i)].value)
                    template.cell(row=row, column=col + 2, value=odoo_datos_dc67['D' + str(j)].value)
                    template.cell(row=row, column=col + 3, value='676000')
                    template.cell(row=row, column=col + 4, value=odoo_datos_bn['E' + str(i)].value)
                    template.cell(row=row, column=col + 5, value=odoo_datos_bn['G' + str(i)].value)
                    template.cell(row=row, column=col + 6, value=odoo_datos_dc67['E' + str(j)].value)
                    template.cell(row=row, column=col + 7, value=odoo_datos_dc67['G' + str(j)].value)
                    row += 1
                    break
            k = 3
            while (str(odoo_datos_dc77['E' + str(k)].value) != 'None'):
                k += 1
                codCta77 = odoo_datos_dc77['E' + str(k)]
                if codNac.value == codCta77.value:
                    template.cell(row=row, column=col + 1, value=odoo_datos_bn['D' + str(i)].value)
                    template.cell(row=row, column=col + 2, value=odoo_datos_dc77['D' + str(k)].value)
                    template.cell(row=row, column=col + 3, value='776000')
                    template.cell(row=row, column=col + 4, value=odoo_datos_bn['E' + str(i)].value)
                    template.cell(row=row, column=col + 5, value=odoo_datos_bn['G' + str(i)].value)
                    template.cell(row=row, column=col + 6, value=odoo_datos_dc77['E' + str(k)].value)
                    template.cell(row=row, column=col + 7, value=odoo_datos_dc77['H' + str(k)].value)
                    row += 1
                    break

    # Guardo el archivo
    wb_plantilla.save(filename='/uploads_folder/Informe.xlsx')

@app.route('/download', methods=['GET'])
def download_file():
    if request.method == 'GET':
        #TODO: Revisar extensiones y nombres de los archivos exsitentes
        #TODO: Ejecutar el algoritmo
        #TODO:Revisar como se pueden tomar los datos siempre de la primera hoja sin necesidad de ponerle un nombre para referenciarla aquí
        #TODO: Borrar los archivos de entrada

        file_process()
        return send_file('uploads_folder/Informe.xlsx', as_attachment=True, cache_timeout=0)

## on a POST request of data 
@app.route('/upload', methods=['POST'])
def upload_file():
    if request.method == 'POST':

        ### Auth
        user_code = str(request.form.get('psw'))
        
        # Open database
        conn = sqlite3.connect(db_loc)
        cursor = conn.cursor()
        cursor.execute("PRAGMA key='%s'"%app_key)
        
        # Run sql query
        cursor.execute('select * from upload where uploadcode="%s"'%user_code)
        result = cursor.fetchall() 
        
        # close as we are done with it
        conn.close()

        if len(result)==0: 
            # If we do not get a match, send a message
            flash('Not a valid Code')
            return redirect(request.url)
        

        if 'files[]' not in request.files:
            flash('No files found, try again.')
            return redirect(request.url)

        files = request.files.getlist('files[]')

        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join( upload_dest, filename))

        flash('Files uploaded')
        return redirect('/upload')



if __name__ == "__main__":
    print('to upload files navigate to http://0.0.0.0:4000/upload')
    app.run(host='0.0.0.0',port=4000,debug=True,threaded=True)
